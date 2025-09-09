#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AsyuLynx Domain Discovery Tool
Created by: AMN (Improved Version)
"""

import os
import socket
import requests
import whois
import dns.resolver
import random
import string
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timezone, timedelta

# ---------------- Banner ---------------- #
def print_banner():
    banner = r"""
   _____    ______________.___.____ ___.____    _____.___._______  ____  ___
  /  _  \  /   _____/\__  |   |    |   \    |   \__  |   |\      \ \   \/  /
 /  /_\  \ \_____  \  /   |   |    |   /    |    /   |   |/   |   \ \     / 
/    |    \/        \ \____   |    |  /|    |___ \____   /    |    \/     \ 
\____|__  /_______  / / ______|______/ |_______ \/ ______\____|__  /___/\  \
        \/        \/  \/                       \/\/              \/      \_/

        AsyuLynx Domain Discovery
        Created by: AMN
    """
    print(banner)

# ---------------- Port Profiles ---------------- #
CRITICAL_PORTS = [
    21,22,23,25,53,80,110,135,139,143,445,
    1433,1521,2049,3306,3389,5432,5900,8080,8443
]
# Full scan will be range(1,65536)

# ---------------- Utilities ---------------- #
def resolve_dns(domain):
    results = {}
    record_types = ["A", "AAAA", "MX", "NS", "TXT", "CNAME"]
    for rtype in record_types:
        try:
            answers = dns.resolver.resolve(domain, rtype, lifetime=5)
            results[rtype] = [str(rdata) for rdata in answers]
        except Exception:
            results[rtype] = []
    return results

def get_ip_info(ip):
    try:
        url = f"http://ip-api.com/json/{ip}?fields=status,message,country,org,isp,query"
        resp = requests.get(url, timeout=5).json()
        if resp.get("status") == "success":
            return {
                "IP": resp.get("query", ""),
                "Country": resp.get("country", ""),
                "ISP": resp.get("isp", ""),
                "Org": resp.get("org", "")
            }
    except Exception:
        return {}
    return {}

def check_http(domain):
    urls = [f"http://{domain}", f"https://{domain}"]
    for url in urls:
        try:
            resp = requests.get(url, timeout=5)
            return resp.status_code
        except Exception:
            continue
    return None

def classify_http(code):
    if code == 200:
        return "Internet"
    elif code == 403:
        return "Intranet"
    elif code == 503:
        return "Misconfig"
    return "Closed"

def scan_ports(ip, mode="critical"):
    open_ports = []
    ports = range(1, 65536) if mode == "full" else CRITICAL_PORTS

    for port in ports:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(0.5)
            if sock.connect_ex((ip, port)) == 0:
                open_ports.append(port)
            sock.close()
        except Exception:
            pass
    return open_ports

def get_whois(domain):
    try:
        data = whois.whois(domain)
        registrar = data.registrar if hasattr(data, "registrar") else "N/A"
        return registrar
    except Exception:
        return "N/A"

def assess_risk(open_ports):
    if not open_ports:
        return "Low"
    critical_ports = {21,22,23,25,135,139,445,1433,1521,3306,3389,5900}
    if any(p in critical_ports for p in open_ports):
        return "High"
    if 80 in open_ports or 443 in open_ports:
        return "Medium"
    return "Low"

# ---------------- Processing ---------------- #
def process_target(target, scan_mode):
    # Check if IP or domain
    try:
        socket.inet_aton(target)
        ip = target
        domain = None
    except socket.error:
        domain = target
        try:
            ip = socket.gethostbyname(domain)
        except Exception:
            ip = None

    # Collect data
    dns_records = resolve_dns(domain) if domain else {}
    whois_registrar = get_whois(domain) if domain else "N/A"
    ip_info = get_ip_info(ip) if ip else {}
    open_ports = scan_ports(ip, scan_mode) if ip else []
    http_code = check_http(domain) if domain else None
    exposure = classify_http(http_code)
    risk = assess_risk(open_ports)

    cname = ", ".join(dns_records.get("CNAME", [])) if dns_records else ""
    isp_org = ip_info.get("Org", "") or ip_info.get("ISP", "")

    return [
        domain if domain else "N/A",
        ip if ip else "N/A",
        whois_registrar,
        isp_org,
        cname,
        ", ".join(str(p) for p in open_ports) if open_ports else "Closed",
        http_code if http_code else "N/A",
        exposure,
        risk
    ]

# ---------------- Excel Helper ---------------- #
def apply_risk_colors(ws):
    risk_colors = {
        "Low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),   # Green
        "Medium": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"), # Yellow
        "High": PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Red
    }
    for row in ws.iter_rows(min_row=2, max_col=9, max_row=ws.max_row):
        risk_cell = row[8]  # Risk Level column (9th column)
        if risk_cell.value in risk_colors:
            risk_cell.fill = risk_colors[risk_cell.value]

# ---------------- Helpers ---------------- #
def generate_random_string(length=6):
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))

# ---------------- Main ---------------- #
def main():
    print_banner()
    mode = input("[+] Mode? (single/bulk): ").strip().lower()
    scan_choice = input("[+] Scan type? (1) Critical SSH/RDP/DB/etc or (2) Full (all ports): ").strip()
    scan_mode = "full" if scan_choice == "2" else "critical"

    results = []
    if mode == "single":
        target = input("[+] Enter domain or IP: ").strip()
        results.append(process_target(target, scan_mode))

    elif mode == "bulk":
        file_path = input("[+] Enter file path with domains/IPs: ").strip()
        if os.path.exists(file_path):
            with open(file_path, "r") as f:
                for line in f:
                    target = line.strip()
                    if target:
                        print(f"[*] Processing {target} ({scan_mode} mode)...")
                        results.append(process_target(target, scan_mode))
        else:
            print("[-] File not found.")
            return
    else:
        print("[-] Invalid mode.")
        return

    # ---------------- Save to Excel ---------------- #
    script_dir = os.path.dirname(os.path.abspath(__file__))
    random_str = generate_random_string()
    today = datetime.now(timezone(timedelta(hours=7)))  # UTC+7 (Jakarta)
    filename = f"{today.strftime('%Y%m%d')}_domain_discovery_{random_str}.xlsx"
    filepath = os.path.join(script_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Domain Discovery"

    headers = ["Domain", "IP", "Registrar", "ISP/Org", "CNAME", 
               "Open Services", "HTTP Code", "Exposure", "Risk Level"]
    ws.append(headers)

    for row in results:
        ws.append(row)

    apply_risk_colors(ws)
    wb.save(filepath)
    print(f"\n[+] Results saved to {filepath}")

if __name__ == "__main__":
    main()
